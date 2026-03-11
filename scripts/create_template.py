"""
生産計画_マクロ.xlsm 雛形作成スクリプト
"""

import zipfile
import shutil
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

OUTPUT_PATH = "/home/user/koushin_seisan/.worktrees/phase1a/生産計画_マクロ.xlsm"
TEMP_XLSX_PATH = OUTPUT_PATH.replace(".xlsm", "_temp.xlsx")


def set_column_widths(ws, widths: dict):
    """列幅を設定する。widths は {列番号: 幅} の辞書"""
    for col_num, width in widths.items():
        ws.column_dimensions[get_column_letter(col_num)].width = width


def make_header(ws, headers: list):
    """ヘッダー行を太字で書き込む"""
    for col, value in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=value)
        cell.font = Font(bold=True)


def create_settings_sheet(wb):
    ws = wb.active
    ws.title = "設定"

    make_header(ws, ["項目名", "値", "備考"])

    data = [
        ("BHプラン保存フォルダ", r"P:\生産計画\input\ ", "BHプランを置くフォルダ"),
        ("BH計画保存版_V8パス", r"P:\保存版\V8_BH計画保存版.xlsx", ""),
        ("BH計画保存版_V9パス", r"P:\保存版\V9_BH計画保存版.xlsx", ""),
        ("BH計画保存版_V8_KPNo列番号", 13, "保存版のKP-No列番号（実際のファイルで確認すること）"),
        ("BH計画保存版_V9_KPNo列番号", 13, "保存版のKP-No列番号（実際のファイルで確認すること）"),
        ("加工対象シート名", "日程表", "光真システムから出力されたシート名"),
        ("列番号_生産計画No(B列)", 2, ""),
        ("列番号_客先名(C列)", 3, ""),
        ("列番号_機種名(F列)", 6, ""),
        ("列番号_型式(G列)", 7, ""),
        ("列番号_追加仕様(K列)", 11, ""),
        ("列番号_数量(L列)", 12, ""),
        ("列番号_順序指示発行日(M列)", 13, ""),
        ("列番号_光真ss出荷日(N列)", 14, ""),
        ("列番号_KP-No(R列)", 18, ""),
        ("列番号_BH型式TYPE(S列)", 19, ""),
        ("列番号_MODEL(U列)", 21, ""),
        ("列番号_属性(I列)", 9, ""),
        ("列番号_機械品番(H列)", 8, ""),
        ("問い合わせ先メール", "", "オムロン担当者メールアドレス"),
    ]

    for row_num, (item, value, note) in enumerate(data, start=2):
        ws.cell(row=row_num, column=1, value=item)
        ws.cell(row=row_num, column=2, value=value)
        ws.cell(row=row_num, column=3, value=note)

    set_column_widths(ws, {1: 32, 2: 42, 3: 32})


def create_model_map_sheet(wb):
    ws = wb.create_sheet("列対応表")

    make_header(ws, ["機種", "MODEL値", "保存版ファイル設定キー", "備考"])

    data = [
        ("V8", "V8", "BH計画保存版_V8パス", ""),
        ("V9", "V9", "BH計画保存版_V9パス", ""),
        ("メンテV8", "ﾒﾝﾃV8", "BH計画保存版_V8パス", ""),
        ("メンテV9", "ﾒﾝﾃV9", "BH計画保存版_V9パス", ""),
    ]

    for row_num, row_data in enumerate(data, start=2):
        for col_num, value in enumerate(row_data, start=1):
            ws.cell(row=row_num, column=col_num, value=value)

    set_column_widths(ws, {1: 25, 2: 25, 3: 25, 4: 25})


def create_calendar_sheet(wb):
    ws = wb.create_sheet("稼働日カレンダー")

    make_header(ws, ["日付", "自社稼働(○/×)", "KMP稼働(○/×)", "備考"])

    set_column_widths(ws, {1: 15, 2: 18, 3: 18, 4: 20})


def create_log_sheet(wb):
    ws = wb.create_sheet("ログ")

    make_header(ws, ["実行日時", "ステップ", "結果", "メッセージ"])

    set_column_widths(ws, {1: 22, 2: 25, 3: 12, 4: 65})


def convert_xlsx_to_xlsm(xlsx_path: str, xlsm_path: str):
    """
    xlsx を xlsm に変換する。
    xlsm は xlsx と同じ ZIP 構造だが、[Content_Types].xml の ContentType と
    workbook の関係を xlsmRelationship に変更する必要がある。
    """
    shutil.copy(xlsx_path, xlsm_path)

    # ZIPとして開いて [Content_Types].xml を書き換える
    import zipfile
    import re

    with zipfile.ZipFile(xlsm_path, "r") as zin:
        names = zin.namelist()
        contents = {name: zin.read(name) for name in names}

    # [Content_Types].xml を修正
    ct_xml = contents["[Content_Types].xml"].decode("utf-8")

    # workbook の ContentType を xlsm 用に変更
    ct_xml = ct_xml.replace(
        'ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"',
        'ContentType="application/vnd.ms-excel.sheet.macroEnabled.main+xml"'
    )
    # Override の Extension も変更（存在する場合）
    ct_xml = ct_xml.replace(
        'Extension="xlsx"',
        'Extension="xlsm"'
    )

    contents["[Content_Types].xml"] = ct_xml.encode("utf-8")

    # vbaProject.bin が無い場合は最小限の空バイナリを追加（xlsmとして認識させるため）
    # 実際のマクロは後からExcelで追加するので空で可
    if "xl/vbaProject.bin" not in contents:
        # 最小限の OLE compound document ヘッダー（空のVBAプロジェクト相当）
        # Excel は開けるが VBA は空の状態になる
        # Content_Types にも追加
        ct_xml_updated = contents["[Content_Types].xml"].decode("utf-8")
        if "vbaProject.bin" not in ct_xml_updated:
            insert_before = "</Types>"
            vba_override = '<Override PartName="/xl/vbaProject.bin" ContentType="application/vnd.ms-office.activeX+xml"/>'
            # Default で追加
            vba_default = '<Default Extension="bin" ContentType="application/vnd.ms-office.activeX"/>'
            if 'Extension="bin"' not in ct_xml_updated:
                ct_xml_updated = ct_xml_updated.replace(insert_before, vba_default + "\n" + insert_before)
            contents["[Content_Types].xml"] = ct_xml_updated.encode("utf-8")

        # xl/_rels/workbook.xml.rels に vbaProject.bin のリレーションを追加
        wb_rels_key = "xl/_rels/workbook.xml.rels"
        if wb_rels_key in contents:
            wb_rels = contents[wb_rels_key].decode("utf-8")
            if "vbaProject.bin" not in wb_rels:
                wb_rels = wb_rels.replace(
                    "</Relationships>",
                    '<Relationship Id="rId_vba" Type="http://schemas.microsoft.com/office/2006/relationships/vbaProject" Target="vbaProject.bin"/>\n</Relationships>'
                )
                contents[wb_rels_key] = wb_rels.encode("utf-8")

        # 空の vbaProject.bin（最小OLE2ヘッダー）
        # D0 CF 11 E0 A1 B1 1A E1 は OLE2 magic bytes
        ole_header = bytes([
            0xD0, 0xCF, 0x11, 0xE0, 0xA1, 0xB1, 0x1A, 0xE1,
            0x00, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00,
            0x00, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00,
            0x3E, 0x00, 0x03, 0x00, 0xFE, 0xFF, 0x09, 0x00,
            0x06, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00,
            0x00, 0x00, 0x00, 0x00, 0x01, 0x00, 0x00, 0x00,
        ] + [0x00] * 464)  # 512 bytes total (one sector)
        contents["xl/vbaProject.bin"] = ole_header

    # 書き直す
    with zipfile.ZipFile(xlsm_path, "w", zipfile.ZIP_DEFLATED) as zout:
        for name, data in contents.items():
            zout.writestr(name, data)


def main():
    wb = Workbook()

    create_settings_sheet(wb)
    create_model_map_sheet(wb)
    create_calendar_sheet(wb)
    create_log_sheet(wb)

    # まず xlsx として保存
    wb.save(TEMP_XLSX_PATH)
    print(f"一時ファイル保存: {TEMP_XLSX_PATH}")

    # xlsm に変換
    convert_xlsx_to_xlsm(TEMP_XLSX_PATH, OUTPUT_PATH)
    print(f"xlsm 生成完了: {OUTPUT_PATH}")

    # 一時ファイル削除
    os.remove(TEMP_XLSX_PATH)
    print("一時ファイル削除完了")

    # シート確認
    import openpyxl
    wb_check = openpyxl.load_workbook(OUTPUT_PATH, keep_vba=True)
    print("\n=== 生成ファイルのシート名一覧 ===")
    for name in wb_check.sheetnames:
        print(f"  - {name}")
    wb_check.close()

    print("\n完了")


if __name__ == "__main__":
    main()
